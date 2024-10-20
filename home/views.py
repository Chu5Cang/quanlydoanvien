from django.db import transaction
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
#  Sử dụng Q objects từ Django để kết hợp nhiều điều kiện tìm kiếm.
from django.db.models import Q
from .models import Doankhoa, Chidoan,Doanvien,Chucvu
from .forms import DoankhoaForm, ChidoanForm,DoanvienForm
# goi thu vien pandas xay dung chuc nang import/export 
import pandas
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment
from .forms import ImportExcelForm 
from datetime import datetime
# Create your views here.
def index(request):
    return render(request, 'pages/home.html')
def doankhoa(request):
    # Lấy thông tin từ query tìm kiếm
    query = request.GET.get('search')
    # Kiểm tra nếu có query, thì lọc theo tên khoa
    if query:
        #  Sử dụng Q objects từ Django để kết hợp nhiều điều kiện tìm kiếm.
        doan_khoas = Doankhoa.objects.filter(Q(tenDK__icontains=query) | Q(maDK__icontains=query))
    else:
        doan_khoas = Doankhoa.objects.all()

    # Trả về trang doankhoa với danh sách doankhoa đã lọc
    return render(request, 'pages/doankhoa.html', {'doan_khoas': doan_khoas})
def add_doankhoa(request):
    if request.method == 'POST':
        form = DoankhoaForm(request.POST)
        if form.is_valid():
            form.save()  # Lưu vào database
            # thông báo 
            messages.success(request, 'Thêm thành công!')
            return redirect('doankhoa')  # Chuyển hướng về trang danh sách Đoàn Khoa
        else:
            messages.error(request, 'Thêm thất bại, vui lòng kiểm tra lại thông tin, có thể thông tin đã tồn tại.')
            return redirect('add_doankhoa')
    else:
        form = DoankhoaForm()
    return render(request, 'pages/add_doankhoa.html', {'form': form})
# Phan chinh sua doan khoa
def edit_doankhoa(request, maDK):
    # Lấy đối tượng Đoàn Khoa theo maDK
    doan_khoa = Doankhoa.objects.get(maDK=maDK)
    # Kiểm tra nếu phương thức là POST để cập nhật thông tin
    if request.method == 'POST':
        form = DoankhoaForm(request.POST, instance=doan_khoa)  # Dùng instance để cập nhật
        if form.is_valid():
            form.save()
            messages.success(request, 'Thông tin Đoàn Khoa đã được cập nhật thành công.')
            return redirect('doankhoa')  # Chuyển hướng về trang danh sách Đoàn Khoa sau khi sửa
    else:
        form = DoankhoaForm(instance=doan_khoa)  # Hiển thị form với dữ liệu của Đoàn Khoa cần sửa

    # Render template với form và đối tượng
    return render(request, 'pages/edit_doankhoa.html', {'form': form})
# Phan xoa Doan Khoa
def delete_doankhoa(request, maDK):
    doan_khoa = Doankhoa.objects.get(maDK=maDK)
    if request.method == 'POST':
        doan_khoa.delete()
        messages.success(request, 'Xóa thành công Đoàn Khoa!')
        return redirect('doankhoa')
    return render(request, 'pages/doankhoa.html', {'object': doan_khoa})
# Chuc nang export DoanKhoa
def export_doankhoa(request):
    # Lấy dữ liệu từ model Doankhoa
    doan_khoas = Doankhoa.objects.all().values('maDK', 'tenDK')

    if not doan_khoas.exists():
        messages.error(request, "Không có dữ liệu để xuất ra file Excel.")
        return redirect('doankhoa')  # Quay lại trang danh sách

    # Tạo workbook và sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách Đoàn Khoa"

    # Tiêu đề cột
    headers = ['STT', 'Mã Khoa', 'Tên Khoa']

    # Kẻ khung cho ô
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Thêm tiêu đề vào sheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')  # Căn giữa tiêu đề
        cell.border = thin_border  # Kẻ khung cho tiêu đề

    # Thêm dữ liệu vào sheet và điều chỉnh chiều rộng cột
    for row_num, doan_khoa in enumerate(doan_khoas, 2):  # Bắt đầu từ hàng thứ 2
        # Cột STT
        cell_stt = ws.cell(row=row_num, column=1, value=row_num - 1)
        cell_stt.alignment = Alignment(horizontal='center')  # Căn giữa cột số thứ tự
        cell_stt.border = thin_border  # Kẻ khung cho cột STT

        # Cột Mã Khoa
        cell_ma_khoa = ws.cell(row=row_num, column=2, value=doan_khoa['maDK'])
        cell_ma_khoa.border = thin_border  # Kẻ khung cho cột Mã Khoa

        # Cột Tên Khoa
        cell_ten_khoa = ws.cell(row=row_num, column=3, value=doan_khoa['tenDK'])
        cell_ten_khoa.border = thin_border  # Kẻ khung cho cột Tên Khoa

        # Điều chỉnh chiều rộng của cột STT (giảm kích thước)
        ws.column_dimensions['A'].width = 5  # Chiều rộng cụ thể cho cột STT
        # Điều chỉnh chiều rộng cho cột Mã Khoa và Tên Khoa
        ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, len(str(doan_khoa['maDK'])) + 2)  # Cột Mã Khoa
        ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, len(str(doan_khoa['tenDK'])) + 2)  # Cột Tên Khoa

    # Kẻ khung cho tất cả các ô trong bảng
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(doan_khoas) + 1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Tạo response với file excel
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="doankhoa.xlsx"'

    # Xuất file Excel
    wb.save(response)
    return response
# chuc nang import DoanKhoa
def import_doankhoa(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']  # Nhận file từ form

            # Đọc file Excel bằng pandas
            try:
                df = pandas.read_excel(excel_file)

                # Kiểm tra xem file có đúng cấu trúc hay không
                if not all(col in df.columns for col in ['STT', 'Mã Khoa', 'Tên Khoa']):
                    messages.error(request, "File không đúng định dạng. Vui lòng kiểm tra lại.")
                    return redirect('import_doankhoa')

                # Lặp qua các dòng và lưu vào model Doankhoa
                for _, row in df.iterrows():
                    Doankhoa.objects.update_or_create(
                        maDK=row['Mã Khoa'],  # Xác định dựa trên Mã Khoa
                        defaults={'tenDK': row['Tên Khoa']}  # Cập nhật tên khoa
                    )
                messages.success(request, 'Import dữ liệu thành công!')
                return redirect('doankhoa')  # Chuyển hướng về trang danh sách
            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")
                return redirect('import_doankhoa')
    else:
        form = ImportExcelForm()

    return render(request, 'pages/import_doankhoa.html', {'form': form})

def chidoan(request):
    query = request.GET.get('search')
    if query:
        chidoans = Chidoan.objects.filter(Q(tenCD__icontains=query) | Q(maCD__icontains=query))
    else:
        chidoans = Chidoan.objects.all()

    return render(request, 'pages/chidoan.html', {'chidoans': chidoans})

def add_chidoan(request):
    if request.method == 'POST':
        form = ChidoanForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thêm thành công!')
            return redirect('chidoan')
        else:
            messages.error(request, 'Thêm thất bại, vui lòng kiểm tra lại thông tin, có thể thông tin đã tồn tại.')
            return redirect('add_chidoan')
    else:
        form = ChidoanForm()
        doankhoas = Doankhoa.objects.all()  # Lấy tất cả các Đoàn Khoa
    return render(request, 'pages/add_chidoan.html', {'form': form, 'doankhoas': doankhoas})

def edit_chidoan(request, maCD):
    chi_doan = Chidoan.objects.get(maCD=maCD)
    # Kiểm tra nếu phương thức là POST để cập nhật thông tin
    if request.method == 'POST':
        form = ChidoanForm(request.POST, instance=chi_doan)  # Dùng instance để cập nhật
        if form.is_valid():
            form.save()
            messages.success(request, 'Thông tin Chi Đoàn đã được cập nhật thành công.')
            return redirect('chidoan')
    else:
        form = ChidoanForm(instance=chi_doan)

    return render(request, 'pages/edit_chidoan.html', {'form': form})

def delete_chidoan(request, maCD):
    chi_doan = Chidoan.objects.get(maCD=maCD)
    if request.method == 'POST':
        chi_doan.delete()
        messages.success(request, 'Xóa thành công Chi Đoàn!')
        return redirect('chidoan')
    return render(request, 'pages/delete_chidoan.html', {'object': chi_doan})
def export_chidoan(request):
    chidoans = Chidoan.objects.all().values('maCD', 'tenCD', 'maDK')

    if not chidoans.exists():
        messages.error(request, "Không có dữ liệu để xuất ra file Excel.")
        return redirect('chidoan')

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách Chi Đoàn"

    headers = ['STT', 'Mã Chi Đoàn', 'Tên Chi Đoàn', 'Đoàn Khoa']

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_num, chidoan in enumerate(chidoans, 2):
        cell_stt = ws.cell(row=row_num, column=1, value=row_num - 1)
        cell_stt.alignment = Alignment(horizontal='center')
        cell_stt.border = thin_border

        cell_ma_chidoan = ws.cell(row=row_num, column=2, value=chidoan['maCD'])
        cell_ma_chidoan.border = thin_border

        cell_ten_chidoan = ws.cell(row=row_num, column=3, value=chidoan['tenCD'])
        cell_ten_chidoan.border = thin_border

        cell_doankhoa = ws.cell(row=row_num, column=4, value=chidoan['maDK'])
        cell_doankhoa.border = thin_border

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(chidoans) + 1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="chidoan.xlsx"'
    wb.save(response)
    return response

def import_chidoan(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                df = pandas.read_excel(excel_file)

                # Kiểm tra định dạng file
                if not all(col in df.columns for col in ['STT', 'Mã Chi Đoàn', 'Tên Chi Đoàn', 'Đoàn Khoa']):
                    messages.error(request, "File không đúng định dạng. Vui lòng kiểm tra lại.")
                    return redirect('import_chidoan')

                # Sử dụng transaction để đảm bảo dữ liệu nhất quán
                with transaction.atomic():
                    for _, row in df.iterrows():
                        # Kiểm tra Đoàn Khoa có tồn tại hay không
                        ma_dk = row['Đoàn Khoa']
                        try:
                            doankhoa = Doankhoa.objects.get(maDK=ma_dk)
                        except Doankhoa.DoesNotExist:
                            messages.error(request, f"Đoàn Khoa '{ma_dk}' không tồn tại. Vui lòng kiểm tra lại.")
                            return redirect('import_chidoan')

                        # Cập nhật hoặc tạo Chi Đoàn
                        Chidoan.objects.update_or_create(
                            maCD=row['Mã Chi Đoàn'],
                            defaults={
                                'tenCD': row['Tên Chi Đoàn'],
                                'maDK': doankhoa  # Liên kết đến đối tượng Đoàn Khoa
                            }
                        )

                messages.success(request, 'Import dữ liệu thành công!')
                return redirect('chidoan')

            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")
                return redirect('import_chidoan')
    else:
        form = ImportExcelForm()

    return render(request, 'pages/import_chidoan.html', {'form': form})

def doanvien(request):
    query = request.GET.get('search')
    if query:
        # chidoans = Chidoan.objects.filter(Q(tenCD__icontains=query) | Q(maCD__icontains=query))
        doanviens = Doanvien.objects.filter(Q(tenDV__icontains=query) | Q(maDV__icontains=query))
        chidoans = Chidoan.objects.filter(Q(maCD__icontains=query))
    else:
        doanviens = Doanvien.objects.all()
        chidoans = Chidoan.objects.all()
        # chidoans =Chidoan.objects.all()
    return render(request, 'pages/doanvien.html', {'doanviens': doanviens, chidoans: chidoans})
def add_doanvien(request):
    if request.method == "POST":
        form = DoanvienForm(request.POST)
        if form.is_valid():  # Kiểm tra xem form có hợp lệ không
            form.save()  # Lưu đối tượng Doanvien vào cơ sở dữ liệu
            return redirect('doanvien')  # Chuyển hướng về trang danh sách Đoàn Viên
        else:
            messages.error(request, 'Thêm thất bại, vui lòng kiểm tra lại thông tin, có thể thông tin đã tồn tại.')
            return redirect('add_doanvien')
    else:
        form = DoanvienForm()  # Tạo một instance mới của DoanvienForm
        chidoan =Chidoan.objects.all()
        chucvus =Chucvu.objects.all()
    return render(request, 'pages/add_doanvien.html', {'form': form,'chidoan':chidoan,'chucvus':chucvus})  # Hiển thị form
def delete_doanvien(request, maDV):
    doan_vien = Doanvien.objects.get(maDV=maDV)  # Lấy đối tượng DoanVien bằng maDV

    if request.method == 'POST':
        doan_vien.delete()  # Xóa đối tượng
        messages.success(request, 'Xóa thành công Đoàn Viên!')  # Thông báo thành công
        return redirect('doanvien')  # Chuyển hướng về trang danh sách đoàn viên

    return render(request, 'pages/doanvien.html', {'object': doan_vien})  # Hiển thị lại thông tin đối tượng
def edit_doanvien(request, maDV):
    doan_vien = Doanvien.objects.get(maDV=maDV)
    
    # Kiểm tra nếu phương thức là POST để cập nhật thông tin
    if request.method == 'POST':
        form = DoanvienForm(request.POST, instance=doan_vien)  # Dùng instance để cập nhật
        if form.is_valid():
            form.save()
            messages.success(request, 'Thông tin Đoàn Viên đã được cập nhật thành công.')  # Cập nhật thông báo
            return redirect('doanvien')  # Chuyển hướng về danh sách Đoàn Viên
    else:
        form = DoanvienForm(instance=doan_vien)  # Sử dụng DoanvienForm

    return render(request, 'pages/edit_doanvien.html', {'form': form})  # Đảm bảo template đúng

def import_doanvien(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                df = pandas.read_excel(excel_file)

                # Kiểm tra định dạng file
                required_columns = ['STT', 'Mã Đoàn Viên', 'Tên Đoàn Viên', 'Mã Chi Đoàn', 'Giới Tính', 'Ngày Sinh', 'Quê Quán', 'Số điện thoại', 'Ngày Vào Đoàn', 'Chức Vụ']
                if not all(col in df.columns for col in required_columns):
                    messages.error(request, "File không đúng định dạng. Vui lòng kiểm tra lại.")
                    return redirect('import_doanvien')

                # Sử dụng transaction để đảm bảo dữ liệu nhất quán
                with transaction.atomic():
                    for _, row in df.iterrows():
                        ma_cd = row['Mã Chi Đoàn']
                        try:
                            chidoan = Chidoan.objects.get(maCD=ma_cd)
                        except Chidoan.DoesNotExist:
                            messages.error(request, f"Chi Đoàn '{ma_cd}' không tồn tại. Vui lòng kiểm tra lại.")
                            return redirect('import_doanvien')

                        # Kiểm tra mã chức vụ
                        ma_cv = row['Chức Vụ']
                        try:
                            chuc_vu = Chucvu.objects.get(tenCV=ma_cv)  # Kiểm tra tên chức vụ
                        except Chucvu.DoesNotExist:
                            messages.error(request, f"Chức vụ '{ma_cv}' không tồn tại. Vui lòng kiểm tra lại.")
                            return redirect('import_doanvien')

                        # Tạo Đoàn Viên
                        doanvien = Doanvien(
                            maDV=row['Mã Đoàn Viên'],
                            tenDV=row['Tên Đoàn Viên'],
                            maCD=chidoan,  # Liên kết đến đối tượng Chi Đoàn
                            gioi_tinh=1 if row['Giới Tính'].strip().lower() == 'nam' else 0,
                            ngay_sinh=datetime.strptime(row['Ngày Sinh'], "%d/%m/%Y").date(),
                            que_quan=row['Quê Quán'],
                            sdt=row['Số điện thoại'],
                            ngay_vao_doan=datetime.strptime(row['Ngày Vào Đoàn'], "%d/%m/%Y").date(),
                            maCV=chuc_vu,  # Liên kết đến đối tượng Chức Vụ
                        )
                        doanvien.save()

                messages.success(request, 'Import dữ liệu thành công!')
                return redirect('doanvien')

            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")
                return redirect('import_doanvien')
    else:
        form = ImportExcelForm()

    return render(request, 'pages/import_doanvien.html', {'form': form})
def export_doanvien(request):
    doanviens = Doanvien.objects.select_related('maCV').values(
        'maDV', 'tenDV', 'maCD', 'gioi_tinh', 'ngay_sinh', 'que_quan', 'sdt', 'ngay_vao_doan', 'maCV__tenCV'
    )

    if not doanviens.exists():
        messages.error(request, "Không có dữ liệu để xuất ra file Excel.")
        return redirect('doanvien')

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách Đoàn Viên"

    headers = ['STT', 'Mã Đoàn Viên', 'Tên Đoàn Viên', 'Mã Chi Đoàn', 'Giới Tính', 'Ngày Sinh', 'Quê Quán', 'Số điện thoại', 'Ngày Vào Đoàn', 'Chức Vụ']

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Thiết lập tiêu đề
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Thêm dữ liệu vào các ô
    for row_num, doanvien in enumerate(doanviens, 2):
        cell_stt = ws.cell(row=row_num, column=1, value=row_num - 1)
        cell_stt.alignment = Alignment(horizontal='center')
        cell_stt.border = thin_border

        cell_ma_doan_vien = ws.cell(row=row_num, column=2, value=doanvien['maDV'])
        cell_ma_doan_vien.border = thin_border

        cell_ten_doan_vien = ws.cell(row=row_num, column=3, value=doanvien['tenDV'])
        cell_ten_doan_vien.border = thin_border

        cell_ma_chidoan = ws.cell(row=row_num, column=4, value=doanvien['maCD'])
        cell_ma_chidoan.border = thin_border

        cell_gioi_tinh = ws.cell(row=row_num, column=5, value='Nam' if doanvien['gioi_tinh'] == 1 else 'Nữ')
        cell_gioi_tinh.border = thin_border

        cell_ngay_sinh = ws.cell(row=row_num, column=6, value=doanvien['ngay_sinh'])
        cell_ngay_sinh.border = thin_border

        cell_que_quan = ws.cell(row=row_num, column=7, value=doanvien['que_quan'])
        cell_que_quan.border = thin_border

        cell_sdt = ws.cell(row=row_num, column=8, value=doanvien['sdt'])
        cell_sdt.border = thin_border

        cell_ngay_vao_doan = ws.cell(row=row_num, column=9, value=doanvien['ngay_vao_doan'])
        cell_ngay_vao_doan.border = thin_border

        # Hiển thị tên chức vụ
        cell_ten_chuc_vu = ws.cell(row=row_num, column=10, value=doanvien['maCV__tenCV'])
        cell_ten_chuc_vu.border = thin_border

    # Thiết lập đường viền cho tất cả các ô
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(doanviens) + 1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Tạo phản hồi HTTP để tải về file Excel
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="doanvien.xlsx"'
    wb.save(response)
    return response
